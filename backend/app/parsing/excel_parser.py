"""
Underwriting Excel model parser.

Every shop lays its model out differently, so instead of hard-coding
absolute cell references (A1-style) this uses label-adjacent scanning:
for each metric we care about, search every cell in the workbook for text
that matches one of that metric's label patterns, then read the value out
of the cell immediately to the right (same row) or, failing that, directly
below it. That mirrors how a human skims a model -- find the row label,
read the number next to it -- and is far more robust to a firm's specific
layout than fixed coordinates.

If a specific firm's model layout is known, pass a `cell_map` (dict of
key -> "SheetName!A1") to skip the heuristic and read exact cells instead --
that is the "map the model's specific cell layout" hook the spec calls for.

Scenario sheets: any sheet whose name matches the downside/bear pattern is
scanned into a second result set, so a model with a Base Case tab and a
Downside tab produces two comparable KPI sets for slide 2 / slide 2b.

Memo Charts sheet: if the workbook has a tab named "Memo Charts" (or "Memo
Chart"), a value found there for a given field wins over the same field found
anywhere else -- that tab is meant to hold the deck's own numbers, so it's a
more reliable source than an underwriting tab where the same label can appear
more than once or sit beside an unrelated figure. It only overrides fields it
actually has a value for, though: a field the Memo Charts tab is silent on
still falls back to the whole-workbook scan rather than disappearing from the
deck.
"""
import re
from datetime import datetime

import openpyxl

LABEL_PATTERNS: dict[str, list[str]] = {
    "purchase_price": [r"purchase price", r"acquisition price", r"total purchase price"],
    "purchase_price_per_unit": [r"price\s*/\s*unit", r"price\s*per\s*unit", r"price\s*/\s*sf", r"price\s*per\s*sf"],
    "peak_cost": [r"peak cost", r"peak basis"],
    "exit_price": [r"exit price", r"gross sale price", r"disposition price"],
    "going_in_cap": [r"going[- ]in cap", r"entry cap rate", r"year 1 cap rate", r"purchase cap rate"],
    "market_cap": [r"market cap rate", r"cap (?:rate )?on market", r"yield on market"],
    "exit_cap": [r"exit cap", r"terminal cap", r"reversion cap"],
    "leverage": [r"leverage", r"ltv", r"loan[- ]to[- ]value"],
    "debt_rate": [r"interest rate", r"debt rate", r"spread", r"coupon"],
    "gross_debt_proceeds": [r"gross debt proceeds", r"debt proceeds", r"loan amount"],
    "initial_equity": [r"initial equity", r"equity required", r"total equity invested"],
    "peak_equity": [r"peak equity", r"max(?:imum)? equity"],
    "hold_period": [r"hold period", r"holding period", r"investment period"],
    "unlevered_irr": [r"unlevered irr", r"unleveraged\.? irr"],
    "levered_irr": [r"(?<!un)levered irr", r"(?<!un)leveraged\.? irr"],
    "unlevered_em": [r"unlevered equity multiple", r"unlevered em\b", r"unlevered cfx"],
    "levered_em": [r"(?<!un)levered equity multiple", r"(?<!un)levered em\b", r"(?<!un)levered cfx"],
    "cash_on_cash": [r"cash[- ]on[- ]cash", r"\bcoc\b"],
    "sf_or_units": [r"total (?:rentable )?(?:square feet|sf|rsf)", r"unit count", r"number of units",
                    r"units at acquisition"],
    "occupancy": [r"current occupancy", r"^occupancy$", r"leased %", r"% leased"],
    "occupancy_at_exit": [r"occupancy at exit", r"exit occupancy"],
    "walt": [r"walt", r"weighted average lease term"],
    # --- Sources & Uses at Close ---
    "total_sources": [r"total sources"],
    "dd_closing_costs": [r"dd\s*/?\s*closing costs", r"closing costs", r"due diligence costs"],
    "working_capital": [r"working capital"],
    "equity_subtotal": [r"equity subtotal"],
    "financing_cost": [r"financing cost", r"loan fees?"],
    "total_uses": [r"total uses"],
    # --- Assumptions ---
    "lease_term_assumption": [r"lease term assumption", r"average lease term"],
    "downtime_assumption": [r"downtime assumption", r"months? vacant", r"downtime \(months\)"],
    "exit_assumption": [r"exit assumption", r"sale assumption", r"disposition assumption"],
}

DOWNSIDE_SHEET_RE = re.compile(r"downside|bear|stress|sensitivity", re.IGNORECASE)

# Some firm models carry a dedicated summary tab (e.g. "Memo Charts") that
# already holds the exact figures the offering-summary exhibit needs, laid
# out cleanly for that purpose -- as opposed to the underwriting tabs, where
# the same label can appear multiple times or sit next to unrelated numbers
# and trip up the label-adjacent scan (e.g. an LTV row's dollar figure getting
# read as "Total Sources"). When a sheet like that exists, its values win over
# the same fields found elsewhere; fields it doesn't have still come from the
# whole-workbook scan.
MEMO_CHARTS_SHEET_RE = re.compile(r"memo\s*charts?", re.IGNORECASE)

# A "$ Per Unit" / "$ PSF" column header in the model. When present, each
# labelled row's value is also read from that column into "{key}_per_unit", so
# per-unit figures on the deck always come from a cell the model actually
# states -- they are never derived by dividing by a unit count.
PER_UNIT_HEADER_RE = re.compile(r"\$?\s*per\s*(?:unit|sf)\b|\bpsf\b|\$\s*/\s*(?:unit|sf)\b", re.IGNORECASE)

# Keys whose per-unit companion the exhibit tables display.
PER_UNIT_KEYS = {
    "purchase_price", "peak_cost", "exit_price", "initial_equity", "peak_equity",
    "gross_debt_proceeds", "dd_closing_costs", "working_capital", "equity_subtotal",
    "financing_cost", "total_sources", "total_uses",
}

PERCENT_KEYS = {
    "going_in_cap", "market_cap", "exit_cap", "leverage", "debt_rate",
    "unlevered_irr", "levered_irr", "cash_on_cash", "occupancy", "occupancy_at_exit",
}
CURRENCY_KEYS = {
    "purchase_price", "purchase_price_per_unit", "peak_cost", "exit_price",
    "gross_debt_proceeds", "initial_equity", "peak_equity",
    "total_sources", "dd_closing_costs", "working_capital", "equity_subtotal",
    "financing_cost", "total_uses",
} | {f"{k}_per_unit" for k in PER_UNIT_KEYS}
MULTIPLE_KEYS = {"unlevered_em", "levered_em"}


SF_LABEL_RE = re.compile(r"square feet|\bsf\b|\brsf\b|\bpsf\b", re.IGNORECASE)


def _normalize(key: str, raw, label: str = "") -> str | None:
    """Format a raw cell value for display.

    `label` is the model's own row label, used only where the same field can
    mean different things depending on how it was written (a size row is
    square feet or unit count depending on its wording).
    """
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    if isinstance(raw, (int, float)):
        if key in PERCENT_KEYS:
            # Excel stores 8% as 0.08; also tolerate models that store "8" literally.
            value = raw * 100 if abs(raw) <= 1 else raw
            # Trim a pointless trailing zero so occupancy reads "97%" while a
            # cap rate still reads "5.1%", matching the reference decks.
            return f"{value:.1f}%".replace(".0%", "%")
        if key in CURRENCY_KEYS:
            return f"${raw:,.0f}"
        if key in MULTIPLE_KEYS:
            return f"{raw:.2f}x"
        if key == "sf_or_units":
            unit = "SF" if SF_LABEL_RE.search(label) else "Units"
            return f"{raw:,.0f} {unit}"
        return str(raw)
    text = str(raw).strip()
    return text or None


def _compiled_patterns() -> dict[str, list[re.Pattern]]:
    return {key: [re.compile(p, re.IGNORECASE) for p in pats] for key, pats in LABEL_PATTERNS.items()}


# Column headers and section captions that sit between a row label and its
# value in real firm layouts. Never a value.
COLUMN_HEADER_RE = re.compile(
    r"^(?:gross|total|net|\$\s*per\s*unit|\$\s*psf|per\s*unit|psf|value|annual rent|"
    r"pricing|cap rate|gross returns|debt|equity|sources? & uses.*|transaction overview)$",
    re.IGNORECASE,
)

NUMERIC_KEYS = PERCENT_KEYS | CURRENCY_KEYS | MULTIPLE_KEYS

# How far right of a label to look for its value. Firm models put single
# values several columns over (e.g. label in col A, value in col C).
VALUE_SCAN_COLS = 6


def _is_label_like(text: str, patterns: dict[str, list[re.Pattern]]) -> bool:
    """True if a candidate cell is really another row label or a column header.

    Without this, a label whose own value cell is empty would swallow the next
    label as its value -- e.g. "Year 1 Cap Rate" picking up the text
    "Market Cap Rate" from the row below.
    """
    stripped = text.strip()
    if not stripped:
        return True
    if COLUMN_HEADER_RE.match(stripped) or PER_UNIT_HEADER_RE.search(stripped):
        return True
    return any(r.search(stripped) for regs in patterns.values() for r in regs)


def _is_plausible(key: str, value: float) -> bool:
    """Reject a numeric cell that can't be the field it was matched to.

    Firm models put several figures on one row -- an LTV row typically carries
    both "65.0%" and the gross debt dollars. Taking the first numeric to the
    right of the label would read the dollar amount as the percentage and
    render "26650000% LTV", so ratio-shaped fields get a sanity range and the
    scan moves on to the next column when a candidate falls outside it.
    """
    if key in PERCENT_KEYS:
        # Either a fraction (0.65) or whole percent (65). Occupancy and
        # cash-on-cash can exceed 100, so allow generous headroom.
        return abs(value) <= 300
    if key in MULTIPLE_KEYS:
        return abs(value) <= 100  # equity multiples live around 1-5x
    return True


def _pick_value(ws, row: int, col: int, key: str, patterns: dict[str, list[re.Pattern]]):
    """Find the value belonging to a label at (row, col).

    Scans rightward along the row first (the dominant layout), then tries the
    cell directly below for vertically-stacked label/value pairs. Numeric
    fields only accept genuinely numeric cells, which keeps stray text out of
    money and percentage columns.
    """
    numeric_only = key in NUMERIC_KEYS

    candidates = [ws.cell(row=row, column=c).value for c in range(col + 1, col + 1 + VALUE_SCAN_COLS)]
    candidates.append(ws.cell(row=row + 1, column=col).value)

    for raw in candidates:
        if raw is None or raw == "":
            continue
        if isinstance(raw, bool):
            continue
        if isinstance(raw, str):
            if numeric_only or _is_label_like(raw, patterns):
                continue
        elif isinstance(raw, (int, float)) and not _is_plausible(key, raw):
            continue
        return raw
    return None


def _find_per_unit_columns(ws, max_row: int, max_col: int) -> list[int]:
    """Column indices of any "$ Per Unit" / "$ PSF" header cells in the sheet.

    Returned in left-to-right order. A model can have several (one per table
    block), so a labelled row reads whichever of them holds a value.
    """
    cols: list[int] = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            value = ws.cell(row=row, column=col).value
            if isinstance(value, str) and PER_UNIT_HEADER_RE.search(value.strip()):
                if col not in cols:
                    cols.append(col)
    return sorted(cols)


def _scan_sheet(ws, patterns: dict[str, list[re.Pattern]]) -> dict:
    found: dict = {}
    max_row = min(ws.max_row, 500)
    max_col = min(ws.max_column, 60)
    per_unit_cols = _find_per_unit_columns(ws, max_row, max_col)

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell.value, str):
                continue
            label = cell.value.strip()
            if not label:
                continue
            for key, regs in patterns.items():
                if key in found:
                    continue
                if any(r.search(label) for r in regs):
                    normalized = _normalize(key, _pick_value(ws, row, col, key, patterns), label)
                    if normalized is not None:
                        found[key] = normalized

                    # Pick up the row's explicitly-stated per-unit figure, if the
                    # sheet has a per-unit column to the right of this label.
                    pu_key = f"{key}_per_unit"
                    if key in PER_UNIT_KEYS and pu_key not in found:
                        for pu_col in per_unit_cols:
                            if pu_col <= col:
                                continue
                            raw = ws.cell(row=row, column=pu_col).value
                            if not isinstance(raw, (int, float)):
                                continue
                            pu_normalized = _normalize(pu_key, raw)
                            if pu_normalized is not None:
                                found[pu_key] = pu_normalized
                                break
    return found


def _scan_workbook(wb, sheet_filter=None) -> dict:
    patterns = _compiled_patterns()
    results: dict = {}
    for ws in wb.worksheets:
        if sheet_filter is not None and not sheet_filter(ws.title):
            continue
        sheet_results = _scan_sheet(ws, patterns)
        for key, value in sheet_results.items():
            results.setdefault(key, value)
    return results


def extract_facts(xlsx_path: str, cell_map: dict | None = None) -> dict:
    """Returns (base_case_facts, downside_case_facts)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if cell_map:
        base_facts = {}
        for key, ref in cell_map.items():
            sheet_name, coord = ref.split("!")
            ws = wb[sheet_name]
            base_facts[key] = _normalize(key, ws[coord].value)
        wb.close()
        return base_facts, {}

    memo_sheet_names = [ws.title for ws in wb.worksheets if MEMO_CHARTS_SHEET_RE.search(ws.title)]
    has_named_scenarios = any(DOWNSIDE_SHEET_RE.search(ws.title) for ws in wb.worksheets)

    if has_named_scenarios:
        whole_workbook_facts = _scan_workbook(wb, sheet_filter=lambda name: not DOWNSIDE_SHEET_RE.search(name))
    else:
        whole_workbook_facts = _scan_workbook(wb)

    if memo_sheet_names:
        # The Memo Charts tab is the more reliable source when it has a
        # figure, but it doesn't necessarily carry every field the exhibit
        # needs -- fields it's silent on still come from the whole-workbook
        # scan rather than disappearing from the deck.
        memo_facts = _scan_workbook(wb, sheet_filter=lambda name: name in memo_sheet_names)
        base_facts = {**whole_workbook_facts, **memo_facts}
    else:
        base_facts = whole_workbook_facts

    downside_facts = (
        _scan_workbook(wb, sheet_filter=lambda name: bool(DOWNSIDE_SHEET_RE.search(name)))
        if has_named_scenarios else {}
    )

    wb.close()
    return base_facts, downside_facts
