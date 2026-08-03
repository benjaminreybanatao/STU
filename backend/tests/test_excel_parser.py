"""Excel model parser regression coverage.

Run with: python -m pytest tests/ -q     (from the backend/ directory)
"""
from pathlib import Path

import openpyxl

from app.parsing.excel_parser import extract_facts


def test_memo_charts_sheet_takes_priority_over_the_whole_workbook_scan(tmp_path):
    """A "Memo Charts" tab is meant to hold the deck's own numbers -- when one
    exists, a field it has should win over the same field found on an
    underwriting tab, which can pick up an unrelated figure sitting next
    to the same label."""
    wb = openpyxl.Workbook()
    underwriting = wb.active
    underwriting.title = "Underwriting"
    underwriting["A1"] = "Total Sources"
    underwriting["B1"] = 200000  # a decoy figure from an unrelated adjacent cell

    memo = wb.create_sheet("Memo Charts")
    memo["A1"] = "Total Sources"
    memo["B1"] = 42065000

    path = tmp_path / "model.xlsx"
    wb.save(str(path))

    base_facts, _ = extract_facts(str(path))
    assert base_facts["total_sources"] == "$42,065,000"


def test_memo_charts_sheet_does_not_blank_out_fields_it_lacks(tmp_path):
    """A Memo Charts tab rarely carries every field the exhibit needs. Fields
    it's silent on must still come from the whole-workbook scan -- an earlier
    version of this feature scanned the Memo Charts sheet exclusively and
    dropped every other field from the deck entirely."""
    wb = openpyxl.Workbook()
    underwriting = wb.active
    underwriting.title = "Underwriting"
    underwriting["A1"] = "Purchase Price"
    underwriting["B1"] = 41000000
    underwriting["A2"] = "Levered IRR"
    underwriting["B2"] = 0.208

    memo = wb.create_sheet("Memo Charts")
    memo["A1"] = "Total Sources"
    memo["B1"] = 42065000

    path = tmp_path / "model.xlsx"
    wb.save(str(path))

    base_facts, _ = extract_facts(str(path))
    assert base_facts["total_sources"] == "$42,065,000"
    assert base_facts["purchase_price"] == "$41,000,000"
    assert base_facts["levered_irr"] == "20.8%"


def test_no_memo_charts_sheet_falls_back_to_the_whole_workbook_scan(tmp_path):
    """Without a Memo Charts tab, behavior is unchanged: scan every sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Underwriting"
    ws["A1"] = "Total Sources"
    ws["B1"] = 42065000

    path = tmp_path / "model.xlsx"
    wb.save(str(path))

    base_facts, _ = extract_facts(str(path))
    assert base_facts["total_sources"] == "$42,065,000"


def test_long_decimals_are_rounded_for_display(tmp_path):
    """Fields with no dedicated $/%/x formatting (WALT, hold period, lease
    term/downtime/exit assumptions) still come straight out of a cell, which
    can hold a formula result like 2.3333333333333335 -- a real deck's
    narrative once read "...resulting in a 2.3333333333333335 WALT." Round to
    something a reader would actually write by hand."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Underwriting"
    ws["A1"] = "WALT"
    ws["B1"] = 2.3333333333333335
    ws["A2"] = "Hold Period"
    ws["B2"] = 4.0

    path = tmp_path / "model.xlsx"
    wb.save(str(path))

    base_facts, _ = extract_facts(str(path))
    assert base_facts["walt"] == "2.3"
    assert base_facts["hold_period"] == "4"
